"""Forum storage, topic CRUD, comments, likes, and AI drafting.

This module is loaded by backend.runtime into one shared application namespace.
Keep feature code here grouped by responsibility; cross-feature functions remain available
through the runtime during this incremental modularization.
"""

def forum_store_path(username):
    forum_dir = USER_DATA_DIR / "_forum"
    forum_dir.mkdir(parents=True, exist_ok=True)
    return forum_dir / "topics.json"


def load_forum_topics(username):
    path = forum_store_path(username)
    if not path.exists():
        return []
    try:
        data = read_json_lenient(path)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def save_forum_topics(username, topics):
    forum_store_path(username).write_text(json.dumps(topics, ensure_ascii=False, indent=2), encoding="utf-8")


def forum_public_topic(topic, with_comments=False):
    item = dict(topic)
    comments = item.get("comments", [])
    likes = item.get("likes", [])
    item["comment_count"] = len(comments)
    item["like_count"] = len(likes)
    item["view_count"] = int(item.get("view_count", 0) or 0)
    item["heat"] = item["view_count"] + item["like_count"] * 5 + item["comment_count"] * 3
    if not with_comments:
        item.pop("comments", None)
    item.pop("likes", None)
    return item


def forum_list_topics(username):
    topics = sorted(load_forum_topics(username), key=lambda item: item.get("updated_at", ""), reverse=True)
    return {"ok": True, "topics": [forum_public_topic(topic) for topic in topics]}


def forum_get_topic(payload, username):
    topic_id = str(payload.get("id", "") or "").strip()
    topics = load_forum_topics(username)
    for topic in topics:
        if topic.get("id") == topic_id:
            topic["view_count"] = int(topic.get("view_count", 0) or 0) + 1
            save_forum_topics(username, topics)
            return {"ok": True, "topic": forum_public_topic(topic, with_comments=True)}
    raise ValueError("话题不存在")


def forum_create_topic(payload, username):
    title = str(payload.get("title", "") or "").strip()
    body = str(payload.get("body", "") or "").strip()
    source = str(payload.get("source", "") or "user").strip() or "user"
    if not title:
        raise ValueError("请填写话题标题")
    if not body:
        raise ValueError("请填写话题内容")
    now = datetime.now().isoformat(timespec="seconds")
    user = find_user(username) or {}
    topic = {
        "id": secrets.token_urlsafe(10),
        "title": title,
        "body": body,
        "source": source,
        "author": user.get("name") or username,
        "created_at": now,
        "updated_at": now,
        "view_count": 0,
        "likes": [],
        "comments": [],
    }
    topics = load_forum_topics(username)
    topics.append(topic)
    save_forum_topics(username, topics)
    return {"ok": True, "topic": forum_public_topic(topic, with_comments=True)}


def forum_add_comment(payload, username):
    topic_id = str(payload.get("topic_id", "") or "").strip()
    content = str(payload.get("content", "") or "").strip()
    parent_id = str(payload.get("parent_id", "") or "").strip()
    if not topic_id:
        raise ValueError("请选择话题")
    if not content:
        raise ValueError("请填写讨论内容")
    topics = load_forum_topics(username)
    user = find_user(username) or {}
    now = datetime.now().isoformat(timespec="seconds")
    for topic in topics:
        if topic.get("id") == topic_id:
            topic.setdefault("comments", []).append({
                "id": secrets.token_urlsafe(8),
                "parent_id": parent_id,
                "author": user.get("name") or username,
                "content": content,
                "created_at": now,
            })
            topic["updated_at"] = now
            save_forum_topics(username, topics)
            return {"ok": True, "topic": forum_public_topic(topic, with_comments=True)}
    raise ValueError("话题不存在")


def forum_toggle_like(payload, username):
    topic_id = str(payload.get("topic_id", "") or "").strip()
    if not topic_id:
        raise ValueError("请选择话题")
    topics = load_forum_topics(username)
    now = datetime.now().isoformat(timespec="seconds")
    for topic in topics:
        if topic.get("id") == topic_id:
            likes = topic.setdefault("likes", [])
            if username in likes:
                likes.remove(username)
                liked = False
            else:
                likes.append(username)
                liked = True
            topic["updated_at"] = now
            save_forum_topics(username, topics)
            item = forum_public_topic(topic, with_comments=True)
            item["liked"] = liked
            return {"ok": True, "topic": item}
    raise ValueError("话题不存在")


def extract_uploaded_text(file_item):
    name = Path(file_item.get("name", "") or "upload.txt").name
    raw = base64.b64decode(file_item.get("data", "") or "")
    suffix = Path(name).suffix.lower()
    if suffix in {".txt", ".md", ".csv"}:
        return raw.decode("utf-8", errors="ignore")
    if suffix == ".docx":
        try:
            with zipfile.ZipFile(io.BytesIO(raw)) as docx:
                root = ET.fromstring(docx.read("word/document.xml"))
            ns = {"w": W_NS}
            parts = []
            for paragraph in root.findall(".//w:p", ns):
                text = "".join(node.text or "" for node in paragraph.findall(".//w:t", ns)).strip()
                if text:
                    parts.append(text)
            return "\n".join(parts)
        except Exception:
            return ""
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            parts = []
            for ws in wb.worksheets[:3]:
                for row in ws.iter_rows(values_only=True):
                    values = [str(value).strip() for value in row if value not in (None, "")]
                    if values:
                        parts.append(" | ".join(values))
            return "\n".join(parts)
        except Exception:
            return ""
    return raw.decode("utf-8", errors="ignore")


def forum_ai_topic(payload, username):
    seed = str(payload.get("seed", "") or "").strip()
    chat = str(payload.get("chat", "") or "").strip()
    files = payload.get("files", []) or []
    document_texts = []
    for item in files[:5]:
        text = extract_uploaded_text(item)
        if text:
            document_texts.append(f"【{Path(item.get('name', '')).name}】\n{text[:6000]}")
    if not seed and not chat and not document_texts:
        raise ValueError("请输入灵感信息、聊天内容，或上传文档")

    settings = assistant_settings()
    if not settings["url"] or not settings["key"]:
        raise ValueError("未配置 AI 接口，请在系统配置中设置 NewAPI 地址和 Key 后再使用智能生成话题")

    content = "\n\n".join([
        f"输入信息：\n{seed}" if seed else "",
        f"聊天内容：\n{chat}" if chat else "",
        f"文档内容：\n{chr(10).join(document_texts)}" if document_texts else "",
    ]).strip()
    try:
        data = request_json(
            settings["url"] + "/v1/chat/completions",
            settings["key"],
            {
                "model": settings["model"],
                "messages": [
                    {
                        "role": "system",
                        "content": (
                            "请根据输入信息，为团队金点子论坛生成一个适合当天讨论的话题。"
                            "话题要具体、有讨论价值，能启发大家围绕工作改进、产品创新、效率提升或项目机会发表观点。"
                            "严格返回 JSON：{\"title\":\"话题标题\",\"body\":\"话题介绍和讨论引导\"}，不要返回 Markdown。"
                        ),
                    },
                    {"role": "user", "content": content},
                ],
                "temperature": 0.7,
            },
            "POST",
            60,
        )
        result_text = data["choices"][0]["message"]["content"].strip()
        result_text = re.sub(r"^```(?:json)?|```$", "", result_text, flags=re.I).strip()
        result = json.loads(result_text)
        title = str(result.get("title", "") or "").strip()
        body = str(result.get("body", "") or "").strip()
    except Exception as exc:
        raise ValueError(f"智能生成话题调用失败：{exc}")

    if not title or not body:
        raise ValueError("智能生成话题没有返回完整标题和内容，请检查模型配置")
    return forum_create_topic({"title": title, "body": body, "source": "ai"}, username)


def forum_ai_comment(payload, username):
    topic_id = str(payload.get("topic_id", "") or "").strip()
    if not topic_id:
        raise ValueError("请选择话题")

    topics = load_forum_topics(username)
    topic = next((item for item in topics if item.get("id") == topic_id), None)
    if not topic:
        raise ValueError("话题不存在")

    comments = topic.get("comments", [])
    discussion = "\n".join(
        f"- {comment.get('author', '成员')}：{comment.get('content', '')}"
        for comment in comments[-20:]
    ) or "暂无评论"
    settings = assistant_settings()
    if not settings["url"] or not settings["key"]:
        raise ValueError("未配置 AI 接口，请在系统配置中设置 NewAPI 地址和 Key 后再使用 AI 潜水评论")

    user = find_user(username) or {}
    display_name = user.get("name") or username or "成员"
    try:
        data = request_json(
            settings["url"] + "/v1/chat/completions",
            settings["key"],
            {
                "model": settings["model"],
                "messages": [
                    {
                        "role": "system",
                        "content": (
                            f"你是{display_name}的 AI 潜水员，正在金点子论坛里协助他阅读话题和已有讨论，"
                            "然后以他的助手身份发表一条简短、有启发的评论。"
                            "评论要像团队讨论里的自然发言：可以总结共识、补充一个角度、提出一个可落地追问或下一步建议。"
                            "不要居高临下，不要写长文，不要使用 Markdown 标题，控制在 120-220 字。"
                        ),
                    },
                    {
                        "role": "user",
                        "content": (
                            f"话题：{topic.get('title', '')}\n\n"
                            f"话题说明：\n{topic.get('body', '')}\n\n"
                            f"已有讨论：\n{discussion}"
                        ),
                    },
                ],
                "temperature": 0.65,
            },
            "POST",
            60,
        )
        content = data["choices"][0]["message"]["content"].strip()
    except Exception as exc:
        raise ValueError(f"AI 潜水评论调用失败：{exc}")

    if not content:
        raise ValueError("AI 潜水评论没有返回内容，请检查模型配置")

    return forum_add_comment({"topic_id": topic_id, "content": content}, f"{display_name}的 AI 潜水员")


# ===== 每日资讯 =====
